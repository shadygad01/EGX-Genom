"""ExcelSeriesCollector: a generic column-mapped collector for macro/price
series published as .xlsx bulletins (the Excel-publishing pattern common to
government statistics agencies, e.g. CAPMAS/MoF-style releases).

Generalizes exactly the `StooqPriceCollector`/`FredCsvCollector` pattern --
"a small adapter, not a new framework" -- to spreadsheets: the collector
itself is generic; a concrete source is just a `column_map` configuration
(which header names are which value/date, which series id each value column
becomes) once its real download endpoint is verified. No such source is
wired to a live URL yet (endpoint verification for CAPMAS/MoF is pending,
per `docs/DATA_ACQUISITION.md`) -- this class exists so wiring one in, once
verified, is a configuration, not new parsing code.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from agx_research.collectors.archive import RawArchive
from agx_research.collectors.base import CollectionBatch, Collector
from agx_research.collectors.raw import RawDocument, build_binary_raw_document
from agx_research.data.schemas import MacroObservation
from agx_research.sources.spec import SourceSpec


class ExcelSeriesCollector(Collector):
    name = "ExcelSeriesCollector"
    version = "1.0.0"

    def __init__(
        self,
        spec: SourceSpec,
        url: str,
        *,
        date_column: str,
        value_columns: dict[str, str],  # {header: series_id}
        archive: RawArchive,
        sheet_name: str | None = None,
        header_row: int = 1,
        fetcher=None,
    ):
        super().__init__(spec, fetcher)
        self.url = url
        self.date_column = date_column
        self.value_columns = value_columns
        self.archive = archive
        self.sheet_name = sheet_name
        self.header_row = header_row

    def fetch(self) -> list[RawDocument]:
        content = self.fetcher.fetch_bytes(self.url, self.spec)
        return [
            build_binary_raw_document(
                source_id=self.spec.id,
                collector=self.name,
                collector_version=self.version,
                original_url=self.url,
                content=content,
                schema_version=self.spec.schema_version,
                license=self.spec.license,
                archive=self.archive,
            )
        ]

    def parse(self, document: RawDocument) -> CollectionBatch:
        content = self.archive.retrieve(document.content_hash)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook[self.sheet_name] if self.sheet_name else workbook.active

        rows = sheet.iter_rows(values_only=True)
        for _ in range(self.header_row):
            header = next(rows)
        header = [str(h).strip() if h is not None else "" for h in header]

        warnings: list[str] = []
        if self.date_column not in header or not any(c in header for c in self.value_columns):
            return CollectionBatch(
                source_id=self.spec.id,
                raw_document_id=document.id,
                parse_warnings=[f"Unexpected header: {header}"],
            )

        date_idx = header.index(self.date_column)
        value_idx = {col: header.index(col) for col in self.value_columns if col in header}

        observations: list[MacroObservation] = []
        for row in rows:
            if row is None or row[date_idx] is None:
                continue
            raw_date = row[date_idx]
            try:
                observation_date = (
                    raw_date if isinstance(raw_date, date) else datetime.fromisoformat(str(raw_date)).date()
                )
            except ValueError:
                warnings.append(f"Unparseable date: {raw_date!r}")
                continue
            for col, idx in value_idx.items():
                cell = row[idx]
                if cell is None:
                    continue
                try:
                    value = float(cell)
                except (TypeError, ValueError):
                    warnings.append(f"Unparseable value in column {col!r}: {cell!r}")
                    continue
                observations.append(
                    MacroObservation(
                        series_id=self.value_columns[col],
                        observation_date=observation_date,
                        value=value,
                    )
                )

        return CollectionBatch(
            source_id=self.spec.id,
            raw_document_id=document.id,
            macro_observations=observations,
            parse_warnings=warnings,
        )
